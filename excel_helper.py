#!/usr/bin/env python3

import pandas as pd
import os
import json
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelCredentialsExporter:
    """Enhanced Excel export functionality for IAM credentials"""
    
    def __init__(self):
        self.output_dir = 'aws/iam/output'
        os.makedirs(self.output_dir, exist_ok=True)
    
    def export_from_json(self, json_file_path: str, excel_filename: str = None) -> str:
        """Export credentials from JSON file to Excel with specific column order"""
        
        if not excel_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_filename = f"iam_credentials_{timestamp}.xlsx"
        
        # Read JSON file
        try:
            with open(json_file_path, 'r') as f:
                data = json.load(f)
        except FileNotFoundError:
            raise FileNotFoundError(f"JSON file not found: {json_file_path}")
        except json.JSONDecodeError:
            raise ValueError(f"Invalid JSON file: {json_file_path}")
        
        # Extract credentials data
        credentials_list = self._extract_credentials_from_json(data)
        
        # Create Excel file
        excel_path = self._create_formatted_excel(credentials_list, excel_filename)
        
        return excel_path
    
    def _extract_credentials_from_json(self, data: Dict[str, Any]) -> List[Dict[str, str]]:
        """Extract credentials from JSON structure and format for Excel with correct column order"""
        credentials_list = []
        
        # Handle the nested structure: data['accounts'][account_name]['users']
        if 'accounts' in data:
            for account_name, account_data in data['accounts'].items():
                if 'users' in account_data:
                    for user in account_data['users']:
                        # Create credential row with exact column names and order as specified
                        credential_row = {
                            'firstname': user.get('real_user', {}).get('first_name', ''),
                            'lastname': user.get('real_user', {}).get('last_name', ''),
                            'mail id': user.get('real_user', {}).get('email', ''),
                            'username': user.get('username', ''),
                            'password': user.get('console_password', ''),
                            'loginurl': user.get('console_url', ''),
                            'homeregion': user.get('region', ''),
                            'accesskey': user.get('access_key_id', ''),
                            'secretkey': user.get('secret_access_key', ''),
                            # Additional fields for internal tracking (not in Excel)
                            '_account': account_name,
                            '_account_id': account_data.get('account_id', '')
                        }
                        credentials_list.append(credential_row)
        
        return credentials_list
    
    def _create_formatted_excel(self, credentials_list: List[Dict[str, str]], filename: str) -> str:
        """Create formatted Excel file with exact column order"""
        
        # Define the EXACT column order as specified by user
        columns = [
            'firstname',
            'lastname', 
            'mail id',
            'username',
            'password',
            'loginurl',
            'homeregion',
            'accesskey',
            'secretkey'
        ]
        
        # Create DataFrame with specific column order
        df = pd.DataFrame(credentials_list)
        
        # Ensure all required columns exist and reorder them
        for col in columns:
            if col not in df.columns:
                df[col] = ''
        
        # Select only the specified columns in the exact order
        df = df[columns]
        
        # Create Excel file with formatting
        filepath = os.path.join(self.output_dir, filename)
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "IAM Credentials"
        
        # Define header styling
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers with styling
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title.upper()  # Make headers uppercase
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border
        
        # Define data cell styling
        data_alignment = Alignment(horizontal="left", vertical="center")
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add data with styling
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = str(value) if value is not None else ""
                cell.alignment = data_alignment
                cell.border = data_border
                
                # Add alternating row colors for better readability
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Set specific column widths for better readability
        column_widths = {
            'A': 15,  # firstname
            'B': 15,  # lastname
            'C': 35,  # mail id
            'D': 25,  # username
            'E': 15,  # password
            'F': 55,  # loginurl
            'G': 15,  # homeregion
            'H': 25,  # accesskey
            'I': 45   # secretkey
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Add autofilter to headers
        ws.auto_filter.ref = f"A1:{chr(64 + len(columns))}1"
        
        # Freeze the header row
        ws.freeze_panes = "A2"
        
        # Save the workbook
        wb.save(filepath)
        
        return filepath
    
    def create_summary_sheet(self, json_file_path: str, excel_filename: str = None) -> str:
        """Create a summary Excel with multiple sheets"""
        
        if not excel_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_filename = f"iam_summary_{timestamp}.xlsx"
        
        # Read JSON file
        with open(json_file_path, 'r') as f:
            data = json.load(f)
        
        filepath = os.path.join(self.output_dir, excel_filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Main credentials sheet with correct column order
            credentials_list = self._extract_credentials_from_json(data)
            columns = [
                'firstname', 'lastname', 'mail id', 'username', 'password',
                'loginurl', 'homeregion', 'accesskey', 'secretkey'
            ]
            
            df_credentials = pd.DataFrame(credentials_list)
            
            # Ensure all columns exist and reorder
            for col in columns:
                if col not in df_credentials.columns:
                    df_credentials[col] = ''
            df_credentials = df_credentials[columns]
            
            # Write credentials sheet
            df_credentials.to_excel(writer, sheet_name='Credentials', index=False)
            
            # Format the credentials sheet
            workbook = writer.book
            credentials_ws = writer.sheets['Credentials']
            
            # Apply header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in credentials_ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths
            column_widths = [15, 15, 35, 25, 15, 55, 15, 25, 45]
            for i, width in enumerate(column_widths, 1):
                credentials_ws.column_dimensions[chr(64 + i)].width = width
            
            # Summary sheet
            summary_data = []
            if 'accounts' in data:
                for account_name, account_data in data['accounts'].items():
                    user_count = len(account_data.get('users', []))
                    summary_data.append({
                        'Account Name': account_name,
                        'Account ID': account_data.get('account_id', ''),
                        'Account Email': account_data.get('account_email', ''),
                        'Users Created': user_count
                    })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Metadata sheet
            metadata = {
                'Created Date': data.get('created_date', ''),
                'Created Time': data.get('created_time', ''),
                'Created By': data.get('created_by', ''),
                'Total Users': data.get('total_users', 0),
                'Export Date': datetime.now().strftime('%Y-%m-%d'),
                'Export Time': datetime.now().strftime('%H:%M:%S'),
                'Excel Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            df_metadata = pd.DataFrame(list(metadata.items()), columns=['Property', 'Value'])
            df_metadata.to_excel(writer, sheet_name='Metadata', index=False)
        
        return filepath

    def create_simple_excel(self, json_file_path: str) -> str:
        """Create a simple Excel file with just credentials in the correct order"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"iam_credentials_{timestamp}.xlsx"
        
        # Read JSON file
        with open(json_file_path, 'r') as f:
            data = json.load(f)
        
        # Extract credentials
        credentials_list = self._extract_credentials_from_json(data)
        
        # Create simple DataFrame
        columns = [
            'firstname', 'lastname', 'mail id', 'username', 'password',
            'loginurl', 'homeregion', 'accesskey', 'secretkey'
        ]
        
        df = pd.DataFrame(credentials_list)
        
        # Ensure all columns exist and reorder
        for col in columns:
            if col not in df.columns:
                df[col] = ''
        df = df[columns]
        
        # Save to Excel
        filepath = os.path.join(self.output_dir, excel_filename)
        df.to_excel(filepath, index=False, sheet_name='Credentials')
        
        return filepath

# Example usage and testing
if __name__ == "__main__":
    exporter = ExcelCredentialsExporter()
    
    # Example JSON structure for testing
    sample_data = {
        "created_date": "2025-06-01",
        "created_time": "17:24:36 UTC",
        "created_by": "varadharajaan",
        "total_users": 2,
        "accounts": {
            "account01": {
                "account_id": "676004093025",
                "account_email": "varadharajaan@proton.me",
                "users": [
                    {
                        "username": "account01_clouduser01",
                        "real_user": {
                            "first_name": "Yongli",
                            "last_name": "Wang",
                            "email": "Yongli.Wang@bakerhughes.com"
                        },
                        "region": "us-east-1",
                        "access_key_id": "",
                        "secret_access_key": "",
                        "console_password": "DemoUser@123",
                        "console_url": "https://676004093025.signin.aws.amazon.com/console"
                    },
                    {
                        "username": "account01_clouduser02",
                        "real_user": {
                            "first_name": "Ruslan",
                            "last_name": "Khokhlov",
                            "email": "Ruslan.Khokhlov@bakerhughes.com"
                        },
                        "region": "us-east-2",
                        "access_key_id": "",
                        "secret_access_key": "",
                        "console_password": "DemoUser@123",
                        "console_url": "https://676004093025.signin.aws.amazon.com/console"
                    }
                ]
            }
        }
    }
    
    # Save sample data for testing
    with open('sample_credentials.json', 'w') as f:
        json.dump(sample_data, f, indent=2)
    
    # Test Excel export
    excel_path = exporter.export_from_json('sample_credentials.json')
    print(f"Excel file created: {excel_path}")
    print("Columns will be in order: firstname, lastname, mail id, username, password, loginurl, homeregion, accesskey, secretkey")